from copy import copy
import json

from openpyxl import load_workbook
from openpyxl.styles import PatternFill



INPUT_FILES = {
    "Python": "cadl-ranch-coverage-python.json",
    "Typescript": "cadl-ranch-coverage-typescript.json",
}

COLUMNS = {
    "Python": 3,
    "Typescript": 4,
}

def test():
    wb = load_workbook('/tmp/excelcolor/color.xlsx')
    ws=wb.worksheets[0]
    c = ws['B2']
    c.internal_value  #

    c.fill = PatternFill("solid", fgColor="DDDDDD")

    wb.save(filename = '/tmp/excelcolor/color2.xlsx')


def do_it():

    # Prepare the data
    tests : dict[str, list]= {}
    for language, file in INPUT_FILES.items():

        # Load the test file
        with open(file, "r") as fd:
            json_data = json.load(fd)

        for test, status in json_data.items():
            test_statuses = tests.setdefault(test, [])
            test_statuses.append((language, status))

    # Load the excel
    wb = load_workbook('cadl_tests_report_template.xlsx')
    ws=wb.worksheets[0]

    # Load the default colors from tab 2
    models=wb.worksheets[1]
    pass_cell = models['A1']
    fail_cell = models['A2']
    not_impl_cell = models['A3']

    # Prepare the Excel
    for language, col_num in COLUMNS.items():
        ws.cell(1, col_num, language)

    current_row = 2
    current_column = 2

    for test, results in tests.items():
        ws.cell(current_row, 2, test)

        for (language, status) in results:

            status_cell = ws.cell(current_row, COLUMNS[language], status)
            if status == 'pass':
                status_cell.fill = copy(pass_cell.fill)
            elif status == 'fail':
                status_cell.fill = copy(fail_cell.fill)
            elif status == 'not-implemented':
                status_cell.fill = copy(not_impl_cell.fill)

        current_row += 1

    # Save it
    ws.page_setup.fitToWidth = 1
    ws.column_dimensions['B'].width = 62.86 # Pragmatic
    ws.column_dimensions['C'].width = 16.29 # Pragmatic
    ws.column_dimensions['D'].width = 16.29 # Pragmatic
    wb.save(filename = 'cadl_tests_report.xlsx')



if __name__ == "__main__":
    do_it()
